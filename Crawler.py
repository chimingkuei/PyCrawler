from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from bs4 import BeautifulSoup
import socket
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import rows_from_range

class Crawler:
    def __init__(self, excel_path, ouputexcel_path):
        self.excel_path = excel_path
        self.ouputexcel_path = ouputexcel_path

    def CheckInternet(self):
        try:
            socket.create_connection(("www.google.com", 80))
            return True
        except OSError:
            pass
        return False

    def ReadExcel(self, path):
            data = pd.read_excel(path)
            return data['公司']

    def find_empty_row(self, ws, column_index):
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, column_index).value is None:
                return row
        return ws.max_row + 1
    
    def WriteExcel(self, data):
        wb = load_workbook(self.ouputexcel_path)
        # 指定要写入的工作表
        sheet_name = 'sheet1'
        ws = wb[sheet_name]
        # 找到要写入的起始行
        start_row = self.find_empty_row(ws, 1)
        #print(start_row)
        # 将数据写入DataFrame
        df = pd.DataFrame(data, index=[0])
        # 将DataFrame写入Excel文件
        with pd.ExcelWriter(self.ouputexcel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet_name, startrow=start_row-1, index=False, header=False)
     
    def catch_element(self, soup, element):
        result = soup.find("td", string=element)
        if result is not None:
            data = result.find_next_sibling("td").get_text(strip=True)
        else:
            data = ""
        return data

    def GrabCompanyInfo(self, company):
        if self.CheckInternet():
            # chrome_options = Options()
            # chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            # chrome_options.add_argument("--disable-gpu")
            # chrome_options.add_argument("--disable-extensions")
            # chrome_options.add_argument("--disable-dev-shm-usage")
            # chrome_options.add_argument("--no-sandbox")
            # driver = webdriver.Chrome(options=chrome_options)
            driver = webdriver.Chrome()
            driver.get("https://findbiz.nat.gov.tw/fts/query/QueryBar/queryInit.do")
            search_box = driver.find_element(By.ID, "qryCond")
            search_box.send_keys(company)
            search_box.send_keys(Keys.RETURN)
            time.sleep(3)
            try:
                search_link = driver.find_element(By.LINK_TEXT, company)
                search_link.click()
                time.sleep(3)
                html_content = driver.page_source
                driver.close()
                # 儲存整個網頁
                # with open('saved_page.html', 'w', encoding='utf-8') as file:
                #     file.write(html_content)
                soup = BeautifulSoup(html_content, 'html.parser')
                # 印出整個網頁
                #print(soup.prettify())
                # 印出網頁部分資訊
                data=""
                if 	"解散" in self.catch_element(soup, "公司狀況"):
                    company_name= self.catch_element(soup, "公司名稱").split('公司')[0]+"公司"
                    data = {
                        '統一編號': ['解散'],
                        '公司名稱': [company_name]
                    }
                else:
                    VAT_number= self.catch_element(soup, "統一編號").rstrip('訂閱')
                    company_name= self.catch_element(soup, "公司名稱").split('公司')[0]+"公司"
                    capital= self.catch_element(soup, "資本總額(元)")
                    actual_capital= self.catch_element(soup, "實收資本額(元)")
                    representative_name = self.catch_element(soup, "代表人姓名")
                    approved_date_of_establishment = self.catch_element(soup, "核准設立日期")
                    business_information = self.catch_element(soup, "所營事業資料")
                    print("統一編號:", VAT_number)
                    print("公司名稱:", company_name)
                    print("資本總額(元):", capital)
                    print("實收資本額(元):", actual_capital)
                    print("代表人姓名:", representative_name)
                    print("核准設立日期:",  approved_date_of_establishment)
                    print("所營事業資料:", business_information)
                    time.sleep(5)
                    data = {
                        '統一編號': [VAT_number],
                        '公司名稱': [company_name],
                        '核准設立日期': [approved_date_of_establishment],
                        '資本總額(元)': [capital],
                        '實收資本額(元)': [actual_capital],
                        '代表人姓名': [representative_name],
                        '所營事業資料': [business_information]
                    }
                self.WriteExcel(data)
            except:
                data = {
                        '統一編號': ['找不到'],
                        '公司名稱': [company],
                    }
                self.WriteExcel(data)
        else:
            print("請確認網路連線!")

    def BatchGrabCompanyInfo(self):
        data = self.ReadExcel(self.excel_path)
        for index in data:
            self.GrabCompanyInfo(index)


if __name__ == '__main__':
    Object = Crawler("Data.xlsx", "output.xlsx")
    Object.BatchGrabCompanyInfo()
